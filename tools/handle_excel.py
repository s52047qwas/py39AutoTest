
import pprint
from openpyxl import load_workbook

from tools.handle_loging import my_logger
from tools.handle_path import case_data_dir
from tools.handle_ini import conf

class HandleExcel:
    def __init__(self,file_path,sheet_name):
        my_logger.info(msg='用例文件名称:')
        my_logger.info(msg=file_path)
        my_logger.info(msg='sheet_name:')
        my_logger.info(msg=sheet_name)
        #excel文件
        self.file_name=file_path
        #初始化表格对象   实例属性
        self.work_book = load_workbook(filename=self.file_name)
        #初始化sheet对象
        self.sheet = self.work_book[sheet_name]

    #获取表头和所有测试用例数据
    def __get_all_data(self):
        all_datas = list(self.sheet.iter_rows(values_only=True))
        #print(list(all_datas))
        title = all_datas[0] # 获取表头，list切片
        case_datas = all_datas[1:] # 获取测试用例数据，list切片
        my_logger.info(msg='excel读取到的测试用例数据：')
        my_logger.info(msg=case_datas)
        my_logger.info(msg='excel读取到的测试用例表头数据：')
        my_logger.info(msg=title)
        return title,case_datas

    # 数据拼接 (dict)
    def get_case_data_dict(self):
        case_list = []
        title, case_datas = self.__get_all_data()
        for val in case_datas:
            result = dict(zip(title, val))
            case_list.append(result)
        self.__close_excel()
        my_logger.info(msg='测试用例数据拼接完成：')
        my_logger.info(msg=case_list)
        return case_list

    #保存测试结果到excel
    def write_result(self,rows,columns,result=None):
        self.sheet.cell(row=rows,column=columns).value=result
        self.__save_excel()
        self.__close_excel()

    #保存excel
    def __save_excel(self):
        self.work_book.save(filename=self.file_name)

    # excel关闭
    def __close_excel(self):
        self.work_book.close()



if __name__ == '__main__':
    sheet_name = conf.get(section='sheet',option='name')
    cl = HandleExcel(file_path=case_data_dir,sheet_name=sheet_name)
    case_list = cl.get_case_data_dict()
    pprint.pprint(case_list)





